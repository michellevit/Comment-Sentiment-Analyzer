import os
import subprocess
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


def main():
    file_path = create_workbook()
    open_file(file_path)


def create_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Setup"
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 100
    ws['A1'] = "YouTube Video URL:"
    ws['A1'].fill = PatternFill(start_color="FF0000", fill_type="solid")
    ws['A1'].font = Font(color="FFFFFF", bold=True)
    ws['A2'] = "Max Amount of Comments To Analyze:"
    ws['A2'].fill = PatternFill(start_color="FF0000", fill_type="solid")
    ws['A2'].font = Font(color="FFFFFF", bold=True)
    ws['B2'] = "100"
    ws['A5'] = "INSTRUCTIONS"
    ws['A5'].fill = PatternFill(start_color="D0D0D0", fill_type="solid")
    ws['A5'].font = Font(bold=True)
    ws.merge_cells('A6:B6')
    ws.merge_cells('A7:B7')
    ws.merge_cells('A8:B8')
    ws.merge_cells('A9:B9')
    ws.merge_cells('A10:B10')
    ws['A6'] = "Step 1: Enter a YouTube URL into cell B1"
    ws['A7'] = "Step 2: Adjust the B2 cell value to analyze more/less comments (note: comment analysis takes ~3 seconds per comment)"
    ws['A8'] = "Step 3: Save the file"
    ws['A9'] = "Step 4: Exit the file"
    ws['A10'] = "Step 5: Double-click the file 'run_script_2.bat'"
    ws.sheet_view.selection[0].activeCell = 'B1'  
    ws.sheet_view.selection[0].sqref = 'B1'
    wb.save(filename='YouTube-Comment-Analyzer-Setup.xlsx')
    file_path = 'YouTube-Comment-Analyzer-Setup.xlsx'
    return file_path



def open_file(path):
    if os.name == 'nt':  # Windows
        os.startfile(path)
    elif os.name == 'posix':  # macOS, Linux
        subprocess.run(['open', path], check=True)


# Run comment analyzer 1
if __name__ == "__main__":
    main()