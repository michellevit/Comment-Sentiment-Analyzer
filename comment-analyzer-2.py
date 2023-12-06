import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from googleapiclient.discovery import build
from transformers import TFAutoModelForSequenceClassification, RobertaTokenizer, pipeline
from dotenv import load_dotenv
import tensorflow as tf 
import warnings
from langdetect import detect


# Use test data in development mode to avoid going over API call limit
DEVELOPMENT_MODE = False 


# CREDENTIALS (in .env file):
load_dotenv()
yt_api_key = os.getenv("YOUTUBE_API_KEY")


def main():
    try:
        model = TFAutoModelForSequenceClassification.from_pretrained("cardiffnlp/twitter-roberta-base-sentiment")
        tokenizer = RobertaTokenizer.from_pretrained("cardiffnlp/twitter-roberta-base-sentiment")
        classifier = pipeline("sentiment-analysis", model=model, tokenizer=tokenizer)
    except Exception as e:
        print(f"Error loading model or tokenizer: {e}")
        return
    try:
        wb = load_workbook("YouTube-Comment-Analyzer-Setup.xlsx")
    except FileNotFoundError:
        print("File 'YouTube-Comment-Analyzer-Setup.xlsx' not found. Please check the file path.")
        return
    ws = wb.active
    youtube = build("youtube", "v3", developerKey=yt_api_key)
    url = ws["B1"].value
    max_results = int(ws["B2"].value)
    estimated_time = 3 * max_results / 60
    print("-----------------")
    print("-----------------")
    print(f"Your data is currently being analyzed. This will take up to {estimated_time:.2f} minutes.")
    print("-----------------")
    if not isinstance(url, str):
        print("Invalid URL in Excel file. Please check the content of B1 cell.")
        return
    yt_video_id = extract_youtube_id(url)
    if DEVELOPMENT_MODE:
        yt_video_title = url
    else:
        os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'  # Suppress TensorFlow logging
        tf.get_logger().setLevel('ERROR')  # Suppress deprecated function warnings
        warnings.filterwarnings('ignore')
        video_response = youtube.videos().list(
            part="snippet",
            id=yt_video_id
        ).execute()
        yt_video_title = video_response["items"][0]["snippet"]["title"]
    prep_sentiment_ws(wb, yt_video_title, "Positive")
    prep_sentiment_ws(wb, yt_video_title, "Neutral")
    prep_sentiment_ws(wb, yt_video_title, "Negative")
    if DEVELOPMENT_MODE:
        yt_video_title = url
        get_mock_yt_comments(wb, classifier)
    else:
        get_yt_comments(youtube, yt_video_id, max_results, wb, classifier, token="")
    totals = get_totals(wb)
    total_positive = totals[0]
    total_neutral = totals[1]
    total_negative = totals[2]
    # if not DEVELOPMENT_MODE:
    #     os.remove("YouTube-Comment-Analyzer-Setup.xlsx")
    total_comments = total_positive + total_neutral + total_negative
    if total_comments != 0:
        percent_positive = "{:.1f}%".format(total_positive * 100 / total_comments)
        percent_neutral = "{:.1f}%".format(total_neutral * 100 / total_comments)
        percent_negative = "{:.1f}%".format(total_negative * 100 / total_comments)
    else:
        percent_positive = "0.0%"
        percent_neutral = "0.0%"
        percent_negative = "0.0%"
    prepare_overview_sheet(wb, yt_video_title, totals, percent_positive, percent_neutral, percent_negative)
    print("-----------------")
    print("Total Positive Comments: ", total_positive, " (", percent_positive, ")")
    print("Total Neutral Comments: ", total_neutral, " (", percent_neutral, ")")
    print("Total Negative Comments: ", total_negative, " (", percent_negative, ")")
    print("-----------------")
    print("Analysis complete.")
    print("To view complete results - open file: YouTube-Comment-Analyzer-Complete.xlsx")
    wb.save(filename="YouTube-Comment-Analyzer-Complete.xlsx")
    file_path = 'YouTube-Comment-Analyzer-Complete.xlsx'
    open_file(file_path)


def extract_youtube_id(url):
    regex = r"(?:v=|youtu\.be/)([^&]+)"
    match = re.search(regex, url)
    if match:
        return match.group(1)
    else:
        return "Invalid YouTube URL"


def prep_sentiment_ws(wb, yt_video_title, sheet_type):
    ws = wb.create_sheet("Comments")
    ws.title = f"{sheet_type} Comments"
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 200
    cell_A1 = ws['A1']
    cell_A1.value = "YouTube Video:"
    cell_A1.alignment = Alignment(horizontal="left")
    cell_A1.fill = PatternFill(start_color="FF0000", fill_type="solid")
    cell_A1.font = Font(color="FFFFFF", bold=True)
    ws.merge_cells('B1:E1')
    title_cell = ws['B1']
    title_cell.alignment = Alignment(horizontal="left")
    title_cell.value = yt_video_title
    title_cell.fill = PatternFill(start_color="FF0000", fill_type="solid")
    title_cell.font = Font(color="FFFFFF")
    headers = ["Comment Date", "Sentiment", "Likes", "Username", "Comment"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=header)
        cell.fill = PatternFill(start_color="D0D0D0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if i != 5 else "left") 
        cell.font = Font(bold=True)


def style_row(ws, row_num, sentiment):
    font_colors = {
        "Positive": "008000",
        "Neutral": "787878",
        "Negative": "FF0000"
    }
    font_color = font_colors.get(sentiment, "000000")
    for col_num in range(1, 6):
        cell = ws.cell(row=row_num, column=col_num)
        cell.alignment = Alignment(horizontal="center" if col_num < 5 else "left")
        if col_num == 2:
            cell.font = Font(color=font_color, bold=True)
        else:
            cell.font = Font(color="000000", bold=False)


def get_mock_yt_comments(wb, classifier):
    mock_comments = [
        {"username": "User1", "date": "2021-01-01", "comment": "Great video!", "like_count": 5},
        {"username": "User2", "date": "2021-01-02", "comment": "Very informative.", "like_count": 3},
        {"username": "User3", "date": "2021-01-03", "comment": "Not great.", "like_count": 0},
        {"username": "User4", "date": "2021-01-04", "comment": "It was ok.", "like_count": 0},
        {"username": "User5", "date": "2020-01-04", "comment": "This video in on YouTube.", "like_count": 0},
    ]
    for comment_data in mock_comments:
        sentiment = score_comment(comment_data["comment"], classifier)
        append_comment_to_sheet(wb, comment_data, sentiment)


def get_yt_comments(youtube, yt_video_id, max_results, wb, classifier, token=""):
    try:
        row_number = 3
        while max_results > 0:  # Continue fetching comments until max_results is reached
            # Set maxResults to the minimum of max_results or 100
            request = youtube.commentThreads().list(
                part="snippet",
                videoId=yt_video_id,
                pageToken=token,
                maxResults=min(max_results, 100), 
                order="time", 
                textFormat="plainText"
            )
            response = request.execute()
            for item in response.get("items", []):
                top_comment = item["snippet"]["topLevelComment"]["snippet"]
                comment_text = top_comment["textDisplay"]
                if len(comment_text) > 512:
                    comment_text = comment_text[:512] 
                sentiment = score_comment(comment_text, classifier)
                try:
                    comment_language = detect(comment_text)
                    if comment_language != 'en':
                        continue  # Skip non-English comments
                except:
                    pass 
                comment_data = {
                    "date": top_comment["publishedAt"][:10],
                    "comment": comment_text,
                    "like_count": top_comment["likeCount"],
                    "username": top_comment["authorDisplayName"]
                }
                append_comment_to_sheet(wb, comment_data, sentiment)
                row_number += 1
                max_results -= 1
                if max_results <= 0:
                    break
            token = response.get("nextPageToken", None)
            if not token or max_results <= 0:
                break 
    except Exception as e:
        print(f"Error fetching YouTube comments: {e}")


def append_comment_to_sheet(wb, comment_data, sentiment):
    ws = wb[f'{sentiment} Comments']
    row_num = ws.max_row + 1
    ws.cell(row=row_num, column=1, value=comment_data["date"])
    ws.cell(row=row_num, column=2, value=sentiment)
    ws.cell(row=row_num, column=3, value=comment_data["like_count"])
    ws.cell(row=row_num, column=4, value=comment_data["username"])
    ws.cell(row=row_num, column=5, value=comment_data["comment"])
    style_row(ws, row_num, sentiment)


def score_comment(comment, classifier):
    outputs = classifier(comment)
    comment_score = outputs[0]["label"]
    if comment_score == "LABEL_0":
        return "Negative"
    elif comment_score == "LABEL_1":
        return "Neutral"
    elif comment_score == "LABEL_2":
        return "Positive"
    else:
        return "ERROR: could not classify comment"


def get_totals(wb):
    totals = {'Positive': 0, 'Neutral': 0, 'Negative': 0}
    for sentiment in totals.keys():
        ws = wb[f'{sentiment} Comments']
        totals[sentiment] = max(ws.max_row - 2, 0)
    return [totals['Positive'], totals['Neutral'], totals['Negative']]


def prepare_overview_sheet(wb, yt_video_title, totals, percent_positive, percent_neutral, percent_negative):
    ws = wb["Setup"]
    ws.title = "Overview"
    # Unmerge any merged cells
    merged_cells_ranges = [merged_cell_range.coord for merged_cell_range in ws.merged_cells.ranges]
    for merged_cell_range in merged_cells_ranges:
        ws.unmerge_cells(merged_cell_range)
    # Clear all cells in the sheet
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.value = None
            if cell.has_style:
                cell.font = Font()
                cell.fill = PatternFill()
                cell.alignment = Alignment()
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws['A1'].fill = PatternFill(start_color="FF0000", fill_type="solid")
    ws['A1'] = "YouTube Video:"
    ws['B1'] = yt_video_title
    ws['A2'].fill = None
    ws['A2'].font = Font(color="000000", bold=True)
    ws['A2'] = "Comments Analyzed:"
    ws['B2'].alignment = Alignment(horizontal="center")
    ws['B2'] = totals[0] + totals[1] + totals[2]
    ws['A4'].fill = PatternFill(start_color="008000", fill_type="solid")
    ws['A4'].font = Font(color="FFFFFF", bold=True)
    ws['A4'] = "Positive Comments:"
    ws['B4'].alignment = Alignment(horizontal="center")
    ws['B4'].font = Font(color="008000", bold=True)
    ws['B4'] = totals[0]
    ws['C4'].alignment = Alignment(horizontal="center")
    ws['C4'].font = Font(color="008000", bold=True)
    ws['C4'] = percent_positive
    ws['A5'].fill = PatternFill(start_color="787878", fill_type="solid")
    ws['A5'].font = Font(color="FFFFFF", bold=True)
    ws['A5'] = "Neutral Comments:"
    ws['B5'].alignment = Alignment(horizontal="center")
    ws['B5'].font = Font(color="787878", bold=True)
    ws['B5'] = totals[1]
    ws['C5'].alignment = Alignment(horizontal="center")
    ws['C5'].font = Font(color="787878", bold=True)
    ws['C5'] = percent_neutral
    ws['A6'].fill = PatternFill(start_color="C70039", fill_type="solid")
    ws['A6'].font = Font(color="FFFFFF", bold=True)
    ws['A6'] = "Negative Comments:"
    ws['B6'].alignment = Alignment(horizontal="center")
    ws['B6'].font = Font(color="C70039", bold=True)
    ws['B6'] = totals[2]
    ws['C6'].alignment = Alignment(horizontal="center")
    ws['C6'].font = Font(color="C70039", bold=True)
    ws['C6'] = percent_negative


def open_file(path):
    if os.name == 'nt':  # Windows
        os.startfile(path)
    elif os.name == 'posix':  # macOS, Linux
        subprocess.run(['open', path], check=True)


# Run comment analyzer 2
if __name__ == "__main__":
    main()
